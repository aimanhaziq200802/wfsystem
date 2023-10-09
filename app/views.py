from django.shortcuts import render, redirect, Http404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from openpyxl import Workbook
from django.utils.text import capfirst
from .models import Quotation, Branch, Department, Entity, QuotationAttachment
from authentication.models import UserProfile
from django.core.mail import send_mail

    
@login_required
def all_quotations(request):

    user_quotations = Quotation.objects.all()

    context = {
        'user_quotations': user_quotations
    }
    return render(request, 'all_quotations.html', context)

@login_required
def create_request(request):

    verifierList = User.objects.filter(userprofile__is_verifier=True)
    departments = Department.objects.all()
    branches = Branch.objects.all()
    entities = Entity.objects.all()

    if request.method == 'POST':
        title = request.POST.get('title')
        remarks = request.POST.get('remarks')
        verifier_username = request.POST.get('verifier')
        entity_id = request.POST.get('entity')
        department_id = request.POST.get('department')
        branch_id = request.POST.get('branch')

        # Ensure all mandatory fields are filled
        if not all([title, remarks, verifier_username, entity_id, department_id, branch_id]):
            messages.error(request, "Please ensure all mandatory fields are filled.")
            return render(request, 'create_request.html', {
                'verifierList': verifierList, 
                'departments': departments, 
                'branches': branches, 
                'entities': entities
            })

        try:
            verifier = User.objects.get(username=verifier_username)
        except User.DoesNotExist:
            messages.error(request, f"User {verifier_username} does not exist.")
            return render(request, 'create_request.html', {
                'verifierList': verifierList, 
                'departments': departments, 
                'branches': branches, 
                'entities': entities
            })

        form_data = {
            'title': title,
            'remarks': remarks,
            'entity': Entity.objects.get(id=entity_id),
            'department': Department.objects.get(id=department_id),
            'branch': Branch.objects.get(id=branch_id),
            'supplier': request.POST.get('supplier'),
            'item': request.POST.get('item'),
            'specification': request.POST.get('specification'),
            'quantity': request.POST.get('quantity'),
            'price': request.POST.get('price'),
            'total_amount': request.POST.get('total_amount'),
            'initiator': request.user,
            'verifier': verifier,
            'verify_status': 'PENDING',
            'verify_remarks': "",
            'approve_status': "",
            'approval_remarks': "",
            'final_approval_status': "",
            'final_approval_remarks': "",
            'finance_status': "",
            'finance_remarks': "",
        }

        new_quotation = Quotation(**form_data)

        try:
            uploaded_files = request.FILES.getlist('attachments')
            new_quotation.full_clean()
            new_quotation.save()

            for file in uploaded_files:
                attachment_instance = QuotationAttachment(quotation=new_quotation, file=file)
                attachment_instance.save()

            messages.success(request, "Your request has been successfully created!")
######################################################################################################################                
#Notify the user
            email_subject = "Your Quotation Has Been Successfully Created"
            email_body = f"""
Hi {capfirst(new_quotation.initiator.username)}!,

Your quotation with ID {new_quotation.qid} has been successfully created and is now pending verification.

From,
Workflow System

This is is auto-generated email. Please do not reply to it.
"""
            send_notification_email(new_quotation.initiator.email, email_subject, email_body)

######################################################################################################################                
#Notify the verifer
            email_subject = "New Quotation Awaiting Your Verification"
            email_body = f"""
Hi {capfirst(new_quotation.verifier.username)}!,

A new quotation with ID {new_quotation.qid} is awaiting your verification.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
            send_notification_email(new_quotation.verifier.email, email_subject, email_body)

        except ValidationError as e:
            messages.error(request, e)
            return render(request, 'create_request.html', {
                'verifierList': verifierList, 
                'departments': departments, 
                'branches': branches, 
                'entities': entities
            })

    return render(request, 'create_request.html', {
        'verifierList': verifierList, 
        'departments': departments, 
        'branches': branches, 
        'entities': entities
    })

@login_required
def quotation_list(request):
    quotationList = Quotation.objects.filter(initiator=request.user)
    context = {
        'quotationList': quotationList
    }
    return render(request, 'quotation_list.html', context)

@login_required
def quotation_details(request, qid):
    try:
        q_detail = Quotation.objects.get(qid=qid)
    except Quotation.DoesNotExist:
        raise Http404("Quotation does not exist")
    
    # Check if the current user is not the initiator
    # if q_detail.initiator != request.user:
    #     messages.error(request, "You are not authorized to view this quotation.")
    #     return redirect('/quotation_list')

    context = {
        'quotation': q_detail,
    }
    return render(request, 'quotation_details.html', context)

@login_required
def verification_queue(request):
    quotationList = Quotation.objects.filter(verifier=request.user, verify_status="PENDING")
    context = {
        'quotationList': quotationList
    }
    return render(request, 'verification_queue.html', context)

@login_required
def verification_history(request):
    # Fetch quotations that the current user has either verified or rejected.
    user_quotations = Quotation.objects.filter(
        verifier=request.user,
        verify_status__in=['VERIFIED', 'REJECTED']
    )

    # user_quotations = Quotation.objects.all()

    context = {
        'quotations': user_quotations
    }
    return render(request, 'verification_history.html', context)

@login_required
def quotation_verify(request, qid):
    try:
        q_detail = Quotation.objects.get(qid=qid)
    except Quotation.DoesNotExist:
        raise Http404("Quotation does not exist")

    # Check if the status of the quotation is not 'PENDING'
    if q_detail.verify_status != "PENDING":
        messages.error(request, "This quotation is no longer editable.")
        return redirect('quotation_details', qid=qid)

    is_verifier = q_detail.verifier == request.user

    if request.method == 'POST' and is_verifier:
        new_verify_status = request.POST.get('verify_status')
        approver_username = request.POST.get('approver')

        if approver_username == "null" or not approver_username:  # "null" is a placeholder for 'None' in your form
            q_detail.approver = None
        else:
            try:
                selected_approver = User.objects.get(username=approver_username)
                q_detail.approver = selected_approver
            except User.DoesNotExist:
                messages.error(request, "Selected approver does not exist.")
        
        if new_verify_status:
            q_detail.verify_status = new_verify_status
            q_detail.verify_remarks = request.POST.get('verify_remarks', '').strip()[:50]

            # Check the verification status
            if new_verify_status == 'VERIFIED':
                q_detail.verified_at = timezone.now()
                q_detail.approve_status = "PENDING"
                # Notify the next approver (assuming you have the approver's details as `q_detail.approver`)
######################################################################################################################
#Notify the initiator
                email_subject = "Your Quotation Status Update"
                email_body = f"""
Hi {capfirst(q_detail.initiator.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.verify_status.lower()} by {capfirst(q_detail.verifier.username)}.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                send_notification_email(q_detail.initiator.email, email_subject, email_body)
######################################################################################################################
#Notify the verifier
                email_subject = "Quotation Verification Notification"
                email_body = f"""
Hi {capfirst(q_detail.verifier.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.verify_status.lower()} by you.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                send_notification_email(q_detail.verifier.email, email_subject, email_body)
######################################################################################################################                
#Notify the approver
                email_subject = "New Quotation Awaiting Approval"
                email_body = f"""
Hi {capfirst(q_detail.approver.username)}!,

There's a new quotation with ID {q_detail.qid} waiting for your approval.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                send_notification_email(q_detail.approver.email, email_subject, email_body)


            elif new_verify_status == 'REJECTED':
                q_detail.approver = None
                q_detail.verified_at = timezone.now()
                q_detail.approve_status = ""  
######################################################################################################################
#Notify the user
                email_subject = "Your Quotation Status Update"
                email_body = f"""
Hi {capfirst(q_detail.initiator.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.verify_status.lower()} by {capfirst(q_detail.verifier.username)}.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                send_notification_email(q_detail.initiator.email, email_subject, email_body)
######################################################################################################################
#Notify the verifier
                email_subject = "Quotation Verification Notification"
                email_body = f"""
Hi {capfirst(q_detail.verifier.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.verify_status.lower()} by you.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                send_notification_email(q_detail.verifier.email, email_subject, email_body)

            q_detail.save()
            messages.success(request, "Verification status updated!")

            return redirect('quotation_details', qid=qid)

    approver_list = User.objects.filter(userprofile__is_approver=True)
    is_verified = q_detail.verify_status == 'VERIFIED'

    context = {
        'quotation': q_detail,
        'is_verifier': is_verifier,
        'approver_list': approver_list,
        'is_verified': is_verified
    }
    return render(request, 'quotation_verify.html', context)

@login_required
def approval_queue(request):
    quotationList = Quotation.objects.filter(approver=request.user, approve_status="PENDING")
    context = {
        'quotationList': quotationList
    }
    return render(request, 'approval_queue.html', context)

@login_required
def approval_history(request):
    user_quotations = Quotation.objects.filter(
        approver=request.user,
        approve_status__in=['APPROVED', 'REJECTED']
    )

    context = {
        'quotations': user_quotations
    }
    return render(request, 'approval_history.html', context)

@login_required
def quotation_approve(request, qid):
    try:
        q_detail = Quotation.objects.get(qid=qid)
    except Quotation.DoesNotExist:
        raise Http404("Quotation does not exist")

    # Check if the status of the quotation is not 'PENDING'
    if q_detail.approve_status != "PENDING":
        messages.error(request, "This quotation is no longer editable.")
        return redirect('quotation_details', qid=qid)

    is_approver = q_detail.approver == request.user

    # Fetch users with the respective roles 
    final_approvers = User.objects.filter(userprofile__is_final_approver=True)
    finance_users = User.objects.filter(userprofile__is_finance=True)

    if request.method == 'POST' and is_approver:
        new_approve_status = request.POST.get('approve_status')
        approval_remarks = request.POST.get('approval_remarks', '').strip()[:50]
        
        if new_approve_status == 'APPROVED':
            forward_decision = request.POST.get('forward_to')
            
            q_detail.approval_at = timezone.now()
            q_detail.approve_status = new_approve_status
            q_detail.approval_remarks = approval_remarks
            
            if forward_decision == "final_approver":
                selected_final_approver_username = request.POST.get('final_approver')
                try:
                    selected_final_approver = User.objects.get(username=selected_final_approver_username)
                    q_detail.final_approver = selected_final_approver
                    q_detail.final_approval_status = "PENDING"

######################################################################################################################
#Notify the initiator
                    email_subject = "Your Quotation Status Update"
                    email_body = f"""
Hi {capfirst(q_detail.initiator.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.approve_status.lower()} by {capfirst(q_detail.approver.username)}.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                    send_notification_email(q_detail.initiator.email, email_subject, email_body)     

######################################################################################################################
#Notify the approver
                    email_subject = "Quotation Approval Notification"
                    email_body = f"""
Hi {capfirst(q_detail.approver.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.approve_status.lower()} by you.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                    send_notification_email(q_detail.approver.email, email_subject, email_body)

######################################################################################################################                
#Notify the Final Approver
                    email_subject = "New Quotation Awaiting Your Final Approval"
                    email_body = f"""
Hi {capfirst(selected_final_approver.username)}!,

A quotation with ID {q_detail.qid} is awaiting your final approval.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                    send_notification_email(selected_final_approver.email, email_subject, email_body)                    

                except User.DoesNotExist:
                    messages.error(request, f"Final Approver {selected_final_approver_username} does not exist.")

            elif forward_decision == "finance":
                selected_finance_username = request.POST.get('finance_user')
                try:
                    selected_finance_user = User.objects.get(username=selected_finance_username)
                    q_detail.finance_user = selected_finance_user
                    q_detail.finance_status = "PENDING"
######################################################################################################################
#Notify the initiator
                    email_subject = "Your Quotation Status Update"
                    email_body = f"""
Hi {capfirst(q_detail.initiator.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.approve_status.lower()} by {capfirst(q_detail.approver.username)}.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                    send_notification_email(q_detail.initiator.email, email_subject, email_body)     

######################################################################################################################
#Notify the approver
                    email_subject = "Quotation Approval Notification"
                    email_body = f"""
Hi {capfirst(q_detail.approver.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.approve_status.lower()} by you.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                    send_notification_email(q_detail.approver.email, email_subject, email_body)

######################################################################################################################                
#Notify the finance user
                    email_subject = "New Quotation Awaiting Your Financial Review"
                    email_body = f"""
Hi {capfirst(selected_finance_user.username)}!,

A quotation with ID {q_detail.qid} is awaiting your financial review.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                    send_notification_email(selected_finance_user.email, email_subject, email_body)

                except User.DoesNotExist:
                    messages.error(request, f"Finance User {selected_finance_username} does not exist.")
            
            

            q_detail.save()
            messages.success(request, "Approval status and remarks updated successfully!")
            return redirect('quotation_details', qid=qid)
        elif new_approve_status == 'REJECTED':
            q_detail.approval_at = timezone.now()
            q_detail.approve_status = new_approve_status
            q_detail.approval_remarks = approval_remarks

            q_detail.final_approver = None
            q_detail.final_approval_status = ""
            # q_detail.final_approval_at = ""
            q_detail.final_approval_remarks = ""

            q_detail.finance_user = None
            q_detail.finance_status = ""
            # q_detail.finance_processed_at = ""
            q_detail.finance_remarks = ""
######################################################################################################################                
#Notify the initiator
            email_subject = "Your Quotation Status Update"
            email_body = f"""
Hi {capfirst(q_detail.initiator.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.approve_status.lower()} by {capfirst(q_detail.approver.username)}.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
            send_notification_email(q_detail.initiator.email, email_subject, email_body)
######################################################################################################################
#Notify the Approver
            email_subject = "Quotation Approval Notification"
            email_body = f"""
Hi {capfirst(q_detail.approver.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.approve_status.lower()} by you.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
            send_notification_email(q_detail.approver.email, email_subject, email_body)

        q_detail.save()
        messages.success(request, "Approval status and remarks updated successfully!")
        return redirect('quotation_details', qid=qid)

    context = {
        'quotation': q_detail,
        'is_approver': is_approver,
        'final_approvers': final_approvers,
        'finance_users': finance_users,
    }
    return render(request, 'quotation_approve.html', context)

@login_required
def final_approval_queue(request):
    quotationList = Quotation.objects.filter(final_approver=request.user, final_approval_status="PENDING")
    context = {
        'quotationList': quotationList
    }
    return render(request, 'final_approval_queue.html', context)

@login_required
def final_approval_history(request):
    # Fetch quotations that the current user has either verified or rejected.
    user_quotations = Quotation.objects.filter(
        final_approver=request.user,
        final_approval_status__in=['APPROVED', 'REJECTED']
    )

    # user_quotations = Quotation.objects.all()

    context = {
        'quotations': user_quotations
    }
    return render(request, 'final_approval_history.html', context)

@login_required
def final_approve(request, qid):
    try:
        q_detail = Quotation.objects.get(qid=qid)
    except Quotation.DoesNotExist:
        raise Http404("Quotation does not exist")

    # Check if the status of the final approval is not 'PENDING'
    if q_detail.final_approval_status != "PENDING":
        messages.error(request, "This quotation has already been processed by the final approver.")
        return redirect('quotation_details', qid=qid)

    is_final_approver = request.user.userprofile.is_final_approver

    # Fetch finance users
    finance_users = User.objects.filter(userprofile__is_finance=True)

    if request.method == 'POST' and is_final_approver:
        new_final_approval_status = request.POST.get('approve_status')
        final_approval_remarks = request.POST.get('final_approval_remarks', '').strip()[:100]

        if new_final_approval_status == 'APPROVED':
            q_detail.final_approval_at = timezone.now()
            q_detail.final_approval_status = new_final_approval_status
            q_detail.final_approval_remarks = final_approval_remarks

            selected_finance_username = request.POST.get('finance_user')
            try:
                selected_finance_user = User.objects.get(username=selected_finance_username)
                q_detail.finance_user = selected_finance_user
                q_detail.finance_status = "PENDING"

######################################################################################################################
#Notify the initiator
                email_subject = "Your Quotation Status Update"
                email_body = f"""
Hi {capfirst(q_detail.initiator.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.final_approval_status.lower()} by {capfirst(q_detail.final_approver.username)}.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                send_notification_email(q_detail.initiator.email, email_subject, email_body)     

######################################################################################################################
#Notify the final approver
                email_subject = "Quotation Final Approval Notification"
                email_body = f"""
Hi {capfirst(q_detail.final_approver.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.final_approval_status.lower()} by you.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                send_notification_email(q_detail.final_approver.email, email_subject, email_body)

######################################################################################################################                
#Notify the Finance user
                email_subject = "New Quotation Awaiting Your Financial Review"
                email_body = f"""
Hi {capfirst(selected_finance_user.username)}!,

A quotation with ID {q_detail.qid} is awaiting your final approval.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
                send_notification_email(selected_finance_user.email, email_subject, email_body)                    

            except User.DoesNotExist:
                messages.error(request, f"Finance User {selected_finance_username} does not exist.")
                return redirect('quotation_details', qid=qid)

            q_detail.save()
            messages.success(request, "Final approval status, remarks, and finance user assignment were updated successfully!")
            return redirect('quotation_details', qid=qid)

        elif new_final_approval_status == 'REJECTED':
            q_detail.final_approval_at = timezone.now()
            q_detail.final_approval_status = new_final_approval_status
            q_detail.final_approval_remarks = final_approval_remarks
            q_detail.finance_user = None
            q_detail.finance_status = ""
            # q_detail.finance_processed_at = ""
            q_detail.finance_remarks = ""

######################################################################################################################
#Notify the initiator
            email_subject = "Your Quotation Status Update"
            email_body = f"""
Hi {capfirst(q_detail.initiator.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.final_approval_status.lower()} by {capfirst(q_detail.final_approver.username)}.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
            send_notification_email(q_detail.initiator.email, email_subject, email_body)     

######################################################################################################################
#Notify the final approver
            email_subject = "Quotation Final Approval Notification"
            email_body = f"""
Hi {capfirst(q_detail.final_approver.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.final_approval_status.lower()} by you.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
            send_notification_email(q_detail.final_approver.email, email_subject, email_body)

            q_detail.save()
            messages.success(request, "Final approval status and remarks updated successfully!")
            return redirect('quotation_details', qid=qid)

    context = {
        'quotation': q_detail,
        'is_final_approver': is_final_approver,
        'finance_users': finance_users  # added to the context
    }
    return render(request, 'final_approve.html', context)

@login_required
def finance_queue(request):
    """List of quotations pending finance processing."""
    quotationList = Quotation.objects.filter(finance_user=request.user, finance_status="PENDING")
    context = {
        'quotationList': quotationList
    }
    return render(request, 'finance_queue.html', context)

@login_required
def finance_history(request):
    """List of quotations processed by the current finance user."""
    user_quotations = Quotation.objects.filter(finance_user=request.user, finance_status__in=['APPROVED', 'REJECTED'])
    context = {
        'quotations': user_quotations
    }
    return render(request, 'finance_history.html', context)

@login_required
def finance_approve(request, qid):
    """Approve or reject a quotation as a finance user."""
    try:
        q_detail = Quotation.objects.get(qid=qid)
    except Quotation.DoesNotExist:
        raise Http404("Quotation does not exist")

    # Ensure the quotation is pending finance processing and the current user is the assigned finance user
    if q_detail.finance_status != "PENDING" or q_detail.finance_user != request.user:
        messages.error(request, "This quotation is either not pending for finance processing or you're not authorized to process it.")
        return redirect('quotation_details', qid=qid)

    if request.method == 'POST':
        new_finance_status = request.POST.get('finance_status')
        finance_remarks = request.POST.get('finance_remarks', '').strip()[:50]

        if new_finance_status == 'APPROVED':
            q_detail.finance_reviewed_at = timezone.now()
            q_detail.finance_status = new_finance_status
            q_detail.finance_remarks = finance_remarks

######################################################################################################################
#Notify the initiator
            email_subject = "Your Quotation Status Update"
            email_body = f"""
Hi {capfirst(q_detail.initiator.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.finance_status.lower()} by {capfirst(q_detail.finance_user.username)}.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
            send_notification_email(q_detail.initiator.email, email_subject, email_body)     

######################################################################################################################
#Notify the finance user
            email_subject = "Quotation Finance Update Notification"
            email_body = f"""
Hi {capfirst(q_detail.finance_user.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.finance_status.lower()} by you.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
            send_notification_email(q_detail.finance_user.email, email_subject, email_body)

            q_detail.save()
            messages.success(request, "Quotation approved successfully!")

        elif new_finance_status == 'REJECTED':
            q_detail.finance_reviewed_at = timezone.now()
            q_detail.finance_status = new_finance_status
            q_detail.finance_remarks = finance_remarks

######################################################################################################################
#Notify the initiator
            email_subject = "Your Quotation Status Update"
            email_body = f"""
Hi {capfirst(q_detail.initiator.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.finance_status.lower()} by {capfirst(q_detail.finance_user.username)}.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
            send_notification_email(q_detail.initiator.email, email_subject, email_body)     

######################################################################################################################
#Notify the finance user
            email_subject = "Quotation Finance Update Notification"
            email_body = f"""
Hi {capfirst(q_detail.finance_user.username)}!,

The quotation with ID {q_detail.qid} has been {q_detail.finance_status.lower()} by you.

From,
Workflow System

This is auto-generated email. Please do not reply to it.
"""
            send_notification_email(q_detail.finance_user.email, email_subject, email_body)

            q_detail.save()
            messages.success(request, "Quotation rejected successfully!")
        
        else:
            messages.error(request, "Invalid action.")

        # After processing or rejecting, redirect to the details page
        return redirect('quotation_details', qid=qid)

    context = {
        'quotation': q_detail,
    }
    return render(request, 'finance_approve.html', context)

@login_required
def export_quotation_to_excel(request, qid):
    try:
        q_detail = Quotation.objects.get(qid=qid)
    except Quotation.DoesNotExist:
        raise Http404("Quotation does not exist")

    wb = Workbook()
    ws = wb.active

    headers = {
        'QID': q_detail.qid,
        'Title': q_detail.title,
        'Date Created': q_detail.date.strftime('%d-%m-%Y %H:%M %p'),
        'Attachment': 'Yes' if q_detail.attachments.exists() else 'No',
        'Remarks': q_detail.remarks,
        'Verify Status': q_detail.get_verify_status_display(),
        'Approve Status': q_detail.get_approve_status_display(),
        'Initiator': q_detail.initiator.username,
        'Verifier': q_detail.verifier.username if q_detail.verifier else 'N/A',
        'Approver': q_detail.approver.username if q_detail.approver else 'N/A',
        'Verified At': q_detail.verified_at.strftime('%d-%m-%Y %H:%M %p') if q_detail.verified_at else 'N/A',
        'Approval At': q_detail.approval_at.strftime('%d-%m-%Y %H:%M %p') if q_detail.approval_at else 'N/A',
        'Verify Remarks': q_detail.verify_remarks,
        'Approval Remarks': q_detail.approval_remarks,
        'Department': q_detail.department.name if q_detail.department else 'N/A',
        'Branch': q_detail.branch.name if q_detail.branch else 'N/A',
        'Entity': q_detail.entity.name if q_detail.entity else 'N/A',
        'Supplier': q_detail.supplier,
        'Item': q_detail.item,
        'Specification': q_detail.specification,
        'Quantity': q_detail.quantity,
        'Price': q_detail.price,
        'Total Amount': q_detail.total_amount,
        'Final Approver': q_detail.final_approver.username if q_detail.final_approver else 'N/A',
        'Final Approval Status': q_detail.get_final_approval_status_display(),
        'Final Approval At': q_detail.final_approval_at.strftime('%d-%m-%Y %H:%M %p') if q_detail.final_approval_at else 'N/A',
        'Final Approval Remarks': q_detail.final_approval_remarks,
        'Finance Status': q_detail.get_finance_status_display(),
        'Finance Approval At': q_detail.finance_reviewed_at.strftime('%d-%m-%Y %H:%M %p') if q_detail.finance_reviewed_at else 'N/A',
        'Finance Remarks': q_detail.finance_remarks,
    }

    for index, (header, field_value) in enumerate(headers.items(), start=1):
        ws.cell(row=index, column=1, value=header)

        if header == 'Attachment':
            attachments = q_detail.attachments.all()
            if attachments:
                attachment_urls = [att.file.url for att in attachments]
                ws.cell(row=index, column=2, value=", ".join(attachment_urls))
            else:
                ws.cell(row=index, column=2, value="No Attachments")
        else:
            ws.cell(row=index, column=2, value=str(field_value))


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=quotation_{qid}.xlsx'
    
    wb.save(response)

    return response

def send_notification_email(to_email, subject, message_body):
    from_email = 'spam88360@gmail.com'
    recipient_list = [to_email]
    send_mail(subject, message_body, from_email, recipient_list)

